import argparse
import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter
import io


DESKTOP_DEFAULT = os.path.join(os.environ["USERPROFILE"], "Desktop") if (
    os.name == "nt" and
    os.path.exists(os.path.join(os.environ["USERPROFILE"], "Desktop"))
) else None
TAIL = 31

def syntax_analyzer() -> tuple[str, str]:
    parser = argparse.ArgumentParser(
        description="CLI for translate Excel to TXT"
    )
    parser.add_argument(
        "inputdir",
        help="Folder with excel files"
        )
    parser.add_argument(
        "outputdir",
        nargs="?",
        default=DESKTOP_DEFAULT,
        help="Folder for result txt file"
    )
    parser.add_argument(
        "-n",
        "--name",
        nargs="?",
        default="result",
        help="Name for result file txt"
    )
    args = parser.parse_args()
    if os.path.isdir(args.inputdir) and os.path.isdir(args.outputdir):
        return args.inputdir, args.outputdir, args.name
    sys.exit("No correct path to output directory")


def check_files(path_to_dir: str, type_file: str="xlsx") -> list:
    result = list()
    for excel_file in os.listdir(path_to_dir):
        if excel_file.split(".")[-1] == type_file:
            result.append(os.path.join(path_to_dir, excel_file))
    return result


def read_excel(path_to_file: str) -> str:
    result = list()
    debug = dict()
    wb = openpyxl.load_workbook(path_to_file)
    sheets = wb['Sheet1']
    rows = sheets.max_row
    columns = sheets.max_column
    for col in range(1, columns + 1):
        for row in range(1, rows + 1):
            if not (sheets.cell(column=col, row=row).value is None):
                if get_column_letter(col) not in debug:
                    debug[get_column_letter(col)] = 1
                else:
                    debug[get_column_letter(col)] += 1
                result.append(sheets.cell(column=col, row=row).value[:TAIL])
    sys.stdout.write("%s\n" % str(debug))
    sys.stdout.write("%s\n" % sum(debug.values()))
    return "\n".join(result)


def write_to_txt(path_to_file: str, data: str, name: str) -> None:
    name_file = os.path.join(path_to_file, f"{name}.txt")
    with open(name_file, "w", encoding="utf8") as file:
        file.write(data)


def main() -> None:
    input, output, name = syntax_analyzer()
    excel_files_list = check_files(input)
    buffer = io.StringIO()
    for excel_file in excel_files_list:
        buffer.write(read_excel(excel_file))
    write_to_txt(
        path_to_file=output,
        data=buffer.getvalue(),
        name=name
    )


if __name__ == "__main__":
    sys.stdout.write("start\n")
    main()
    sys.stdout.write("end\n")
